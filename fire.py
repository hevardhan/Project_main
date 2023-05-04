from pyrebase import initialize_app as p_init
from re import search, fullmatch
from firebase_admin import credentials
from firebase_admin import initialize_app as f_admin_init
from firebase_admin import db, auth, _auth_utils
from database import display_date,date
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl import Workbook
from openpyxl.chart import DoughnutChart, BarChart, Reference
import os.path

# Firebase Configurationcand Credentials---------------------------------------------------------------
firebaseConfig = {
  "apiKey": "AIzaSyCpjyclqZqiL5lkrEGH5KZEwyli3NXNJI0",
  "authDomain": "budget-buddy-11.firebaseapp.com",
  "databaseURL": "https://budget-buddy-11-default-rtdb.firebaseio.com",
  "projectId": "budget-buddy-11",
  "storageBucket": "budget-buddy-11.appspot.com",
  "messagingSenderId": "994473005551",
  "appId": "1:994473005551:web:cf14c6d203454ec627661a",
  "measurementId": "G-6EYJ0B0PXS"
}

cred = credentials.Certificate('D:\PROJECTS\Budget Buddy\Project\Project_main\\budget-buddy-key.json')

# Initializing Firebase------------------------------------------------------------------
f_admin_init(cred, {"databaseURL": "https://budget-buddy-11-default-rtdb.firebaseio.com"})


firebase = p_init(firebaseConfig)

authent = firebase.auth()

# Checking Email and Password ----------------------------------------------------------
def check(email, pwd):
    paswd_ok = False
    
    while check_mail(email) == True:
        if len(pwd) < 8:
            paswd_ok = False
            break
        
        elif not search("[a-z]", pwd):
            paswd_ok = False
            break

        elif not search("[A-Z]", pwd):
            paswd_ok = False
            break

        elif not search("[0-9]", pwd):
            paswd_ok = False
            break
        
        elif search("%{,}$[;/]\s", pwd):
            paswd_ok = False
            break

        else:
            paswd_ok = True
            break
    
    return paswd_ok

# Checking Email------------------------------------------------------------------------
def check_mail(email):

    regex = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,7}\b'

    if fullmatch(regex,email):
        return True
    else:
        return False
 
# Check User ID 
def check_userID(userid):
    if userid.islower() == True:
        try:
            auth.get_user(uid=userid)
            return 1
        except auth.UserNotFoundError:
            return 0
    else:
        return 2    

#Create User---------------------------------------------------------------------------------
def create_user(user_id,email,paswd, confirm_paswd, disp_name):
    if check_userID(userid=user_id) == 0:
        if check(email,paswd) == True:
            if paswd == confirm_paswd:
                try:
                    auth.create_user(uid = f"{user_id}", email = f"{email}", password = f"{paswd}", display_name = f"{disp_name}")
                    create_db(user_id=user_id,disp_name=disp_name)
                    return 0
                except:
                    return 1   
            else:
                return 2
        else:
            return 3
    else:
        return 4

#Login --------------------------------------------------------------------------------
def login(user_id, paswd):
    try:
        user_details = auth.get_user(user_id)
    except _auth_utils.UserNotFoundError:
        return 1
    except:
        return 3
    else:
        mail = user_details.email
        try:
            authent.sign_in_with_email_and_password(email = mail,password= paswd)
            return 0
        except :
            return 2

# Create Database-----------------------------------------------------------------------
ref = db.reference("Userdata")
def create_db(user_id, disp_name):
    ref.child(f"{user_id}").set({"Name": f"{disp_name}","Number of Transactions":0})

# Settings Data--------------------------------------------------------------------------
def add_setting_data(user_id, setting_data):    
    ref.child(f"{user_id}").child("Settings_data").set(setting_data)

#Forgot Password-------------------------------------------------------------------------
def forgot_paswd(user_email):
    try:
        user_details = auth.get_user_by_email(user_email)
        authent.send_password_reset_email(user_email)
        return 0
    except:
        return 1

# Add Income / Expense ------------------------------------------------------------------------
def add_expense(user_id, td, date):
    n = ref.child(f"{user_id}").get()
    p = n['Number of Transactions']
    p +=1
    ref.child(f"{user_id}").update({"Number of Transactions":p})
    
    check = ref.child(f"{user_id}").child("Transaction").child(str(date)).get()
    try: 
        t_no = check['Transaction Number']
        t_no +=1
        ref.child(f"{user_id}").child("Transaction").child(str(date)).update({"Transaction Number":f"{t_no}"})
        ref.child(f"{user_id}").child("Transaction").child(str(date)).child(f"{t_no}").set(td)

    except:
        ref.child(f"{user_id}").child("Transaction").child(str(date)).set({"Transaction Number": 1})
        ref.child(f"{user_id}").child("Transaction").child(str(date)).child(f"{1}").set(td)


# Getting Balance ---------------------------------------------------------------------- 
def balance(user_id):
    trans_ref = db.reference(f'Userdata/{user_id}/Transaction')
    trans_details = trans_ref.order_by_key().get()
    
    sum_inc = 0
    sum_exp = 0

    for date in trans_details:
        for key in trans_details[date]:
            if key.isnumeric():
                if trans_details[date][key]['Type'] == "Income":
                    sum_inc += int(trans_details[date][key]['Amount'])
                else:
                    sum_exp += int(trans_details[date][key]['Amount'])

    return sum_inc - sum_exp

# MOnthly Balance --------------------------------------------------
def monthly_balance(userid):
    trans_ref = db.reference(f'Userdata/{userid}/Transaction')
    trans_details = trans_ref.order_by_key().get()
    

    months = ["Jan","Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    bal = [0,0,0,0,0,0,0,0,0,0,0,0]

    for date in trans_details:
        for i in range(1,12):
            if months[i] in date:
                for key in trans_details[date]:
                    if key.isnumeric():
                        if trans_details[date][key]['Type'] == "Income":
                            bal[i] += int(trans_details[date][key]['Amount'])
                        else:
                            bal[i] -= int(trans_details[date][key]['Amount'])
    
    return bal

def monthly_expense(userid):
    trans_ref = db.reference(f'Userdata/{userid}/Transaction')
    trans_details = trans_ref.order_by_key().get()

    months = ["Jan","Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    exp = [0,0,0,0,0,0,0,0,0,0,0,0]

    for date in trans_details:
        for i in range(1,12):
            if months[i] in date:
                for key in trans_details[date]:
                    if key.isnumeric():
                        if trans_details[date][key]['Type'] == "Expense":
                            exp[i] += int(trans_details[date][key]['Amount'])
                        else:
                            pass

    return exp

# Getting Display Name ------------------------------------------------------------------
def get_display_name(uid):
    details = auth.get_user(uid)
    return details.display_name

# History --------------------------------------------------------------------------------
def history_per_day(user_id, date):
    trans_ref = db.reference(f'Userdata/{user_id}/Transaction')
    trans_details = trans_ref.order_by_key().get()
    trans_data = []

    for key in trans_details[date]:
        if key.isnumeric():
            trans_data.append(trans_details[date][key])
            
    return trans_data

def expense_per_day(user_id, date):
    trans_ref = db.reference(f'Userdata/{user_id}/Transaction')
    trans_details = trans_ref.order_by_key().get()

    sum_exp = 0

    for key in trans_details[date]:
        if key.isnumeric():
            if trans_details[date][key]['Type'] == "Expense":
                sum_exp += trans_details[date][key]['Amount']

    return sum_exp


# Export Data ------------------------------------------------------------
def export(user_id, pswd,path):
    trans_ref = db.reference(f'Userdata/{user_id}/Transaction')
    trans_data = trans_ref.order_by_key().get()

    name_ref = db.reference(f'Userdata/{user_id}')
    name_data = name_ref.order_by_key().get()
    name = name_data['Name']
    t_no = name_data['Number of Transactions']

    wb = Workbook()
    ws = wb.active.title("Main")
    ws1 = wb.create_sheet("Expenses")
    ws2 = wb.create_sheet("Monthly Balance")
    ch1 = wb.create_chartsheet("Daily Average")
    ch2 = wb.create_chartsheet("Weekly Average")
    ch3 = wb.create_chartsheet("Monthly Expense")

    # Worksheet -----------------------
    ws.merge_cells('A1:F1')
    cell1 = ws['A1']
    cell1.value = f'Name: {name}'

    ws.merge_cells('A2:F2')
    cell2 = ws['A2']
    cell2.value = f'Number of Transactions : {t_no}'

    ws.merge_cells('A3:F3')
    cell2 = ws['A3']
    cell2.value = f'Balance : {balance(user_id)}'


    ws['A4'] = 'SNO'
    ws['B4'] = 'Date'
    ws['C4'] = 'Description'
    ws['D4'] = 'Category'
    ws['E4'] = 'Amount'
    ws['F4'] = 'Type'
    

    i = 5
    j = 1

    for date in trans_data:
        for key in trans_data[date]:
            if key.isnumeric():
                ws[f'A{i}'] = j
                ws[f'B{i}'] = date
                ws[f'C{i}'] = trans_data[date][key]['Description']
                ws[f'E{i}'] = int(trans_data[date][key]['Amount'])
                ws[f'F{i}'] = trans_data[date][key]['Type']
                try:
                    ws[f'D{i}'] = trans_data[date][key]['Category']
                except:
                    ws[f'D{i}'] = "NA"
                
                i +=1
                j +=1

    # Expenses  -----------------------
    ws1.merge_cells('A1:E1')
    cell1 = ws1['A1']
    cell1.value = f'Expenses'


    ws1['A2'] = 'SNO'
    ws1['B2'] = 'Date'
    ws1['C2'] = 'Description'
    ws1['D2'] = 'Category'
    ws1['E2'] = 'Amount'

    i = 3
    j = 1

    for date in trans_data:
        for key in trans_data[date]:
            if key.isnumeric():
                if trans_data[date][key]['Type'] == 'Expense':
                    ws1[f'A{i}'] = j
                    ws1[f'B{i}'] = date
                    ws1[f'C{i}'] = trans_data[date][key]['Description']
                    ws1[f'E{i}'] = int(trans_data[date][key]['Amount'])               
                    ws1[f'D{i}'] = trans_data[date][key]['Category']
           
                    i +=1
                    j +=1

    # Doughnut Chart --------------------------------------------------
    chart5 = DoughnutChart()

    labels = Reference(ws1, min_col= 4, min_row=3, max_row=i-1)
    data = Reference(ws1, min_col=5, min_row=2, max_row=i-1)

    chart5.add_data(data, titles_from_data=True)
    chart5.set_categories(labels)
    chart5.title = "Expense Distribution"
    chart5.style = 26
    ws.add_chart(chart5, "H1")

    
    # Daily Average Expenses ---------------------------------------------------
    months = ("Jan","Feb", "Mar", "Apr", "May", "Jun","Jul", "Aug", "Sep","Oct", "Nov", "Dec")

    ws1['G1'] = 'Months'
    ws1['H1'] = 'Daily Average'
    ws1['I1'] = 'Weekly Average'
    ws1['J1'] = 'Total Expense'

    expense = monthly_expense(user_id)
    daily_avg = []
    weekly_avg = []

    for exp in expense:
        if exp != 0:
            daily_avg.append(exp/30)
            weekly_avg.append(exp/7)
        else:
            daily_avg.append(0)
            weekly_avg.append(0) 

    for i in range(1,13):
        ws1[f'G{i+1}'] = months[i-1]
        ws1[f'H{i+1}'] = daily_avg[i-1]
        ws1[f'I{i+1}'] = weekly_avg[i-1]
        ws1[f'J{i+1}'] = expense[i-1]


    # Monthly Balances ----------------------------------------------------------------
    ws2.merge_cells('A1:B1')
    cell1 = ws2['A1']
    cell1.value = f'Balance'

    balance_data = monthly_balance(user_id)

    for i in range(1,13):
        ws2[f'A{i+1}'] = months[i-1]
        ws2[f'B{i+1}'] = balance_data[i-1]


    # Monthly Balance Bar Chart ----------------------------------------------------
    chart1 = BarChart()

    labels = Reference(ws2, min_col=1, min_row=2, max_row=13)
    data = Reference(ws2, min_col=2, min_row=1, max_row=13)

    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(labels)
    chart1.title = "Monthly Balances"
    chart1.style = 26
    ws2.add_chart(chart1, "D1")

    # Daily Average Bar Chart -------------------------------------------------------
    chart2 = BarChart()

    labels = Reference(ws1, min_col=7, min_row=2, max_row=13)
    data = Reference(ws1, min_col=8, min_row=1, max_row=13)
    
    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(labels)
    chart2.title = "Daily Average Expense"
    chart2.style = 26
    ch1.add_chart(chart2)

    # Weekly Average Bar Chart ------------------------------------------------------
    chart3 = BarChart()

    data = Reference(ws1, min_col=9, min_row=1, max_row=13)

    chart3.add_data(data, titles_from_data=True)
    chart3.set_categories(labels)
    chart3.title = "Weekly Average Expense"
    chart3.style = 26
    ch2.add_chart(chart3)

    # Monthly Expense Chart ----------------------------------------------------------
    chart4 = BarChart()

    data = Reference(ws1, min_col=10, min_row=1, max_row=13)
    chart4.add_data(data, titles_from_data=True)
    chart4.set_categories(labels)
    chart4.title = "Monthly Expense"
    chart4.style = 26
    ch3.add_chart(chart4)

    

    filename = 'Transactions.xlsx'

    wb.security = WorkbookProtection(workbookPassword=f'{pswd}', lockStructure=True, lockWindows=True)
    ws.protection.sheet = True
    ws.protection.password = f'{pswd}'

    # Getting the user's download folder and saving it there
    wb.save(f'{path}\{filename}')
    downloads = os.path.join(os.path.expanduser('~'), 'Downloads')
    wb.save(f'{path}/{filename}')
          

